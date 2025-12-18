# -*- coding: utf-8 -*-
"""
Скрипт для обработки данных из success_data.xlsx по требованиям:
1. 1 строка - название четверти
2. 2 строка - заголовки (Аты-жөні + предметы)
3. Данные учеников
4. Строки "5", "4", "3" с подсчетом по предметам
5. Столбцы "5", "4", "3" справа с подсчетом по ученикам
6. Строки "Качество" и "Успеваемость" по предметам
7. Строки "Качество по классу" и "Успеваемость по классу"
8. Отступ 2 строки и следующая четверть
"""
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict

# Цветовая палитра
COLORS = {
    'header_bg': '4472C4',  # Синий для заголовков
    'header_text': 'FFFFFF',  # Белый текст
    'quarter_bg': 'D9E1F2',  # Светло-синий для названия четверти
    'quarter_text': '203864',  # Темно-синий текст
    'data_bg': 'FFFFFF',  # Белый для данных
    'stats_5_bg': 'E2EFDA',  # Светло-зеленый для строки "5"
    'stats_4_bg': 'FFF2CC',  # Светло-желтый для строки "4"
    'stats_3_bg': 'FCE4D6',  # Светло-оранжевый для строки "3"
    'quality_bg': 'E2EFDA',  # Светло-зеленый для качества
    'performance_bg': 'FFF2CC',  # Светло-желтый для успеваемости
    'class_stats_bg': 'BDD7EE',  # Светло-голубой для статистики по классу
    'border': 'BFBFBF',  # Серый для границ
}

# Маппинг четвертей на названия
QUARTER_NAMES = {
    'І': '1 четверть',
    'I': '1 четверть',
    'II': '2 четверть',
    'III': '3 четверть',
    'IV': '4 четверть',
    'Ж': 'Годовая'
}

# Порядок четвертей
QUARTERS_ORDER = ['I', 'II', 'III', 'IV', 'Ж']


def normalize_quarter(quarter):
    """Нормализует четверть к стандартному формату"""
    if not quarter:
        return None
    q = str(quarter).strip()
    mapping = {
        'і': 'I', '1': 'I', 'i': 'I',
        'іі': 'II', '2': 'II', 'ii': 'II',
        'ііі': 'III', '3': 'III', 'iii': 'III',
        'іv': 'IV', 'iv': 'IV', '4': 'IV',
        'ж': 'Ж', '5': 'Ж', 'год': 'Ж', 'годовая': 'Ж'
    }
    if q in ['I', 'II', 'III', 'IV', 'Ж', 'І', 'ІІ', 'ІІІ', 'ІV']:
        if q in ['І', 'I', 'i', 'і']:
            return 'I'
        elif q in ['ІІ', 'II', 'ii', 'іі']:
            return 'II'
        elif q in ['ІІІ', 'III', 'iii', 'ііі']:
            return 'III'
        elif q in ['ІV', 'IV', 'iv', 'іv']:
            return 'IV'
        elif q == 'Ж':
            return 'Ж'
    if q.lower() in mapping:
        return mapping[q.lower()]
    return q


def parse_grade(value):
    """Преобразует значение в оценку (число 1-5)"""
    if pd.isna(value) or value == '' or value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            return int(value) if value == int(value) and 1 <= int(value) <= 5 else None
        value_str = re.sub(r'[^\d]', '', str(value).strip())
        if value_str:
            grade = int(value_str)
            if 1 <= grade <= 5:
                return grade
    except:
        pass
    return None


def get_cell_value_safe(ws, row, col):
    """Безопасное получение значения ячейки с учетом объединенных ячеек"""
    # Проверяем объединенные ячейки
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return ws.cell(row, col).value


def read_data_with_two_level_headers(input_file, sheet_name):
    """Читает данные с двухуровневой структурой заголовков"""
    wb = load_workbook(input_file, data_only=True)
    ws = wb[sheet_name]
    
    # Читаем первые 2 строки для заголовков
    max_col = ws.max_column
    headers = []
    subjects_map = {}  # {col_index: (subject_name, quarter)}
    
    # Обрабатываем каждую колонку
    col = 1
    current_subject = None
    
    while col <= max_col:
        val1 = get_cell_value_safe(ws, 1, col)
        val2 = get_cell_value_safe(ws, 2, col)
        
        val1_str = str(val1).strip() if val1 else ''
        val2_str = str(val2).strip() if val2 else ''
        
        # Проверяем, служебная ли колонка (rowspan=2)
        # Первая колонка может быть пустой, вторая - "Аты-жөні"
        is_service = any(x in val1_str.lower() for x in ['параллель', 'номер_строки', 'аты-жөні', 'фио', 'аты'])
        is_service2 = any(x in val2_str.lower() for x in ['параллель', 'номер_строки', 'аты-жөні', 'фио', 'аты'])
        
        # Если обе строки пустые - это пустая колонка (первая колонка без заголовка)
        if not val1_str and not val2_str:
            # Пустая колонка (первая колонка без заголовка)
            headers.append('')
            subjects_map[col] = ('', None)
            col += 1
        elif is_service or is_service2:
            # Служебная колонка (например, "Аты-жөні")
            header_value = val1_str if val1_str else val2_str
            headers.append(header_value)
            subjects_map[col] = (header_value, None)
            col += 1
        elif val1_str and not is_service:
            # Это предмет
            current_subject = val1_str
            if val2_str:
                quarter = normalize_quarter(val2_str)
                if quarter:
                    headers.append(f"{current_subject}_{quarter}")
                    subjects_map[col] = (current_subject, quarter)
                else:
                    headers.append(current_subject)
                    subjects_map[col] = (current_subject, None)
            else:
                headers.append(current_subject)
                subjects_map[col] = (current_subject, None)
            col += 1
        elif val2_str and not val1_str:
            # Продолжение предмета (colspan)
            quarter = normalize_quarter(val2_str)
            if current_subject and quarter:
                headers.append(f"{current_subject}_{quarter}")
                subjects_map[col] = (current_subject, quarter)
            else:
                headers.append(val2_str)
                subjects_map[col] = (None, quarter)
            col += 1
        else:
            # Пропускаем полностью пустые колонки (не создаем заголовок)
            # Не создаем колонки с названием "Column_X" - это лишние колонки
            col += 1
    
    # Читаем данные начиная с 3-й строки
    data_rows = []
    num_headers = len(headers)
    
    for row in range(3, ws.max_row + 1):
        row_data = []
        # Читаем только столько колонок, сколько заголовков
        for col_idx in range(num_headers):
            col = col_idx + 1  # Колонки начинаются с 1
            if col <= max_col:
                val = get_cell_value_safe(ws, row, col)
                row_data.append(val if val is not None else '')
            else:
                row_data.append('')
        
        # Проверяем, что строка не пустая и содержит значимые данные
        # Ищем колонку с ФИО (обычно это вторая колонка после служебных)
        has_fio = False
        has_data = False
        
        # Проверяем первые несколько колонок на наличие ФИО
        for i, val in enumerate(row_data[:min(5, len(row_data))]):
            if val and str(val).strip():
                val_str = str(val).strip().lower()
                # Проверяем, что это не служебное значение
                if not any(x in val_str for x in ['параллель', 'номер', 'column']):
                    # Если это похоже на ФИО (содержит буквы, не только цифры)
                    if any(c.isalpha() for c in val_str) and len(val_str) > 2:
                        has_fio = True
                    has_data = True
                    break
        
        # Если нет ФИО, но есть другие данные, тоже считаем валидной строкой
        if not has_fio:
            has_data = any(v for v in row_data if v and str(v).strip() and 
                          not str(v).strip().lower() in ['параллель', 'номер_строки', 'column'])
        
        # Добавляем строку только если есть данные
        if has_data:
            data_rows.append(row_data)
    
    wb.close()
    
    # Создаем DataFrame с точным соответствием количества колонок
    if data_rows:
        # Убеждаемся, что все строки имеют одинаковую длину
        max_cols_in_data = max(len(row) for row in data_rows)
        num_cols = min(num_headers, max_cols_in_data)
        
        # Обрезаем заголовки и данные до нужного количества
        final_headers = headers[:num_cols]
        final_data = []
        for row in data_rows:
            # Дополняем или обрезаем строку до нужной длины
            if len(row) < num_cols:
                row = row + [''] * (num_cols - len(row))
            else:
                row = row[:num_cols]
            final_data.append(row)
        
        df = pd.DataFrame(final_data, columns=final_headers)
    else:
        df = pd.DataFrame(columns=headers)
    
    return df, subjects_map


def merge_duplicate_columns(df):
    """Объединяет дублирующиеся столбцы (одинаковый предмет+четверть)"""
    # Группируем столбцы по предмету и четверти
    columns_by_subject_quarter = defaultdict(list)
    
    # Список служебных колонок, которые нужно исключить
    service_columns = ['Параллель', 'Номер_строки', 'Аты-жөні', 'Фио', 'FIO']
    
    for col in df.columns:
        col_str = str(col)
        
        # Пропускаем служебные колонки
        if any(service in col_str for service in service_columns):
            continue
        
        # Пропускаем колонки с названием "Column_X" или "Колонка_X" - это лишние колонки
        col_lower = col_str.lower()
        if (col_str.startswith('Column_') and col_str[7:].isdigit()) or \
           (col_str.startswith('Колонка_') and any(c.isdigit() for c in col_str[8:])):
            continue
        
        # Пропускаем колонки, содержащие "Column" или "Колонка" в названии (кроме точного совпадения)
        if ('column' in col_lower or 'колонка' in col_lower) and col_str not in ['Column', 'Колонка']:
            continue
        
        if '_' in col_str:
            parts = col_str.rsplit('_', 1)
            if len(parts) == 2:
                subject = parts[0]
                quarter = normalize_quarter(parts[1])
                # Пропускаем, если предмет содержит "Column" или "Колонка"
                subject_lower = str(subject).lower() if subject else ''
                if 'column' in subject_lower or 'колонка' in subject_lower:
                    continue
                key = (subject, quarter) if quarter else (subject, None)
                columns_by_subject_quarter[key].append(col)
            else:
                # Пропускаем, если название содержит "Column" или "Колонка"
                if ('column' in col_lower or 'колонка' in col_lower) and col_str not in ['Column', 'Колонка']:
                    continue
                columns_by_subject_quarter[(col_str, None)].append(col)
        else:
            # Пропускаем, если название содержит "Column" или "Колонка"
            if ('column' in col_lower or 'колонка' in col_lower) and col_str not in ['Column', 'Колонка']:
                continue
            columns_by_subject_quarter[(col_str, None)].append(col)
    
    # Объединяем столбцы
    merged_data = {}
    for (subject, quarter), cols in columns_by_subject_quarter.items():
        if len(cols) == 1:
            # Если дубликатов нет, просто берем значения
            merged_data[(subject, quarter)] = df[cols[0]].tolist()
        else:
            # Если есть дубликаты (например, для подгрупп)
            print(f"  Объединение столбцов для ({subject}, {quarter}): {cols}")
            # Создаем список для объединенных значений
            merged_values = []
            for idx in df.index:
                # Получаем значения из всех дублирующихся столбцов для этой строки
                # Используем .values для получения numpy массива, избегая проблем с Series
                values_from_duplicate_cols = df.loc[idx, cols].values
                # Фильтруем непустые значения
                # pd.notna возвращает булевую маску, которую можно использовать для фильтрации
                # np.where может помочь найти индексы непустых, но проще отфильтровать значения
                non_empty_values = []
                for val in values_from_duplicate_cols:
                     # Проверяем, не является ли значение NaN и не пустая ли это строка
                     if pd.notna(val):
                         val_str = str(val).strip()
                         if val_str and val_str not in ['nan', 'None', '']:
                             non_empty_values.append(val_str)
                # Или с использованием list comprehension и numpy
                # non_empty_values = [str(v).strip() for v in values_from_duplicate_cols
                #                     if pd.notna(v) and (isinstance(v, str) or pd.notna(v)) and str(v).strip() and str(v).strip() not in ['nan', 'None', '']]

                # Объединяем непустые значения, убирая дубликаты
                if non_empty_values:
                    # Убираем дубликаты, сохраняя порядок
                    unique_values = []
                    seen = set()
                    for val in non_empty_values:
                        if val not in seen:
                            unique_values.append(val)
                            seen.add(val)
                    # Если только одно значение, не добавляем запятую
                    if len(unique_values) == 1:
                        merged_val = unique_values[0]
                    else:
                        merged_val = ', '.join(unique_values)
                else:
                    merged_val = ''
                merged_values.append(merged_val)
            merged_data[(subject, quarter)] = merged_values
    
    return merged_data


def create_quarter_table(ws, start_row, quarter_name, merged_data, fio_column_data):
    """Создает таблицу для одной четверти с сохранением дизайна"""
    current_row = start_row
    
    # Фильтруем данные по четверти, исключая служебные колонки
    quarter_data = {}
    service_subjects = ['Параллель', 'Номер_строки', 'Аты-жөні', 'Фио', 'FIO']
    
    for (subject, quarter), values in merged_data.items():
        # Пропускаем служебные колонки
        if subject and any(service in str(subject) for service in service_subjects):
            continue
        # Пропускаем колонки без предмета
        if not subject or subject is None:
            continue
        # Пропускаем колонки с названием "Column_X" или "Колонка_X" или содержащие "Column"/"Колонка"
        subject_str = str(subject)
        subject_lower = subject_str.lower()
        if (subject_str.startswith('Column_') and subject_str[7:].isdigit()) or \
           (subject_str.startswith('Колонка_') and any(c.isdigit() for c in subject_str[8:])):
            continue
        if ('column' in subject_lower or 'колонка' in subject_lower) and subject_str not in ['Column', 'Колонка']:
            continue
        if quarter == quarter_name:
            quarter_data[subject] = values
    
    # Отладочный вывод
    if quarter_data:
        print(f"    Найдено предметов для четверти {quarter_name}: {len(quarter_data)}")
        print(f"    Предметы: {list(quarter_data.keys())[:5]}")
    
    if not quarter_data:
        return current_row
    
    subjects = list(quarter_data.keys())
    num_students = len(fio_column_data) if fio_column_data else 0
    
    # Определяем размеры таблицы
    # +3: колонка номеров (без названия) + ФИО + предметы
    stats_col_start = len(subjects) + 3
    total_cols = stats_col_start + 3  # +3: столбцы "5", "4", "3"
    
    # Границы
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )
    
    # 1. Строка с названием четверти
    quarter_display = QUARTER_NAMES.get(quarter_name, quarter_name)
    ws.merge_cells(f'A{current_row}:{get_column_letter(total_cols)}{current_row}')
    cell = ws.cell(current_row, 1)
    cell.value = quarter_display
    cell.font = Font(bold=True, size=14, color=COLORS['quarter_text'])
    cell.fill = PatternFill(start_color=COLORS['quarter_bg'], end_color=COLORS['quarter_bg'], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    current_row += 1
    
    # 2. Строка заголовков
    col_idx = 1
    # Первая колонка - порядковые номера (без названия)
    cell = ws.cell(current_row, col_idx)
    cell.value = ''  # Пустое название
    cell.font = Font(bold=True, color=COLORS['header_text'], size=10)
    cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    col_idx += 1
    
    # Вторая колонка - ФИО
    cell = ws.cell(current_row, col_idx)
    cell.value = 'Аты-жөні'
    cell.font = Font(bold=True, color=COLORS['header_text'], size=10)
    cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    col_idx += 1
    
    for subject in subjects:
        cell = ws.cell(current_row, col_idx)
        cell.value = subject
        cell.font = Font(bold=True, color=COLORS['header_text'], size=10)
        cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1
    
    for grade in ['5', '4', '3']:
        cell = ws.cell(current_row, col_idx)
        cell.value = grade
        cell.font = Font(bold=True, color=COLORS['header_text'], size=10)
        cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1
    
    ws.row_dimensions[current_row].height = 40
    current_row += 1
    
    # 3. Данные учеников
    actual_student_count = 0
    for student_idx in range(num_students):
        # Получаем ФИО для проверки
        fio_value = fio_column_data[student_idx] if student_idx < len(fio_column_data) else ''
        
        # Пропускаем строки с пустым ФИО
        if not fio_value or (isinstance(fio_value, str) and not fio_value.strip()):
            continue
        
        col_idx = 1
        actual_student_count += 1
        
        # Первая колонка - порядковый номер (1, 2, 3...)
        cell = ws.cell(current_row, col_idx)
        cell.value = actual_student_count  # Номер начинается с 1
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.fill = PatternFill(start_color=COLORS['data_bg'] if actual_student_count % 2 == 0 else 'F9F9F9', 
                               end_color=COLORS['data_bg'] if actual_student_count % 2 == 0 else 'F9F9F9', 
                               fill_type="solid")
        col_idx += 1
        
        # Вторая колонка - ФИО
        cell = ws.cell(current_row, col_idx)
        cell.value = fio_value
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        cell.fill = PatternFill(start_color=COLORS['data_bg'] if actual_student_count % 2 == 0 else 'F9F9F9', 
                               end_color=COLORS['data_bg'] if actual_student_count % 2 == 0 else 'F9F9F9', 
                               fill_type="solid")
        col_idx += 1
        
        # Оценки по предметам
        student_grades = []
        for subject in subjects:
            values = quarter_data[subject]
            value = values[student_idx] if student_idx < len(values) else ''
            cell = ws.cell(current_row, col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = PatternFill(start_color=COLORS['data_bg'] if actual_student_count % 2 == 0 else 'F9F9F9', 
                                   end_color=COLORS['data_bg'] if actual_student_count % 2 == 0 else 'F9F9F9', 
                                   fill_type="solid")
            
            # Обновляем логику для подсчета оценок, если значение содержит несколько оценок (через запятую)
            if value:
                # Разбиваем строку на части по запятой и пробелу
                parts = str(value).split(',')
                for part in parts:
                    grade = parse_grade(part.strip())
                    if grade:
                        student_grades.append(grade)
            col_idx += 1
        
        # Статистика по ученику
        count_5 = sum(1 for g in student_grades if g == 5)
        count_4 = sum(1 for g in student_grades if g == 4)
        count_3 = sum(1 for g in student_grades if g == 3)
        
        for count, grade_color in [(count_5, 'E2EFDA'), (count_4, 'FFF2CC'), (count_3, 'FCE4D6')]:
            cell = ws.cell(current_row, col_idx)
            cell.value = count
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=grade_color, 
                                   end_color=grade_color, 
                                   fill_type="solid")
            col_idx += 1
        
        current_row += 1
    
    # 4. Строки "5", "4", "3" по предметам
    grade_colors = {'5': COLORS['stats_5_bg'], '4': COLORS['stats_4_bg'], '3': COLORS['stats_3_bg']}
    
    for grade in ['5', '4', '3']:
        col_idx = 1
        # Первая колонка - пустая (для номеров)
        cell = ws.cell(current_row, col_idx)
        cell.value = grade
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=grade_colors[grade], end_color=grade_colors[grade], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col_idx += 1
        
        # Вторая колонка - пустая (для ФИО)
        cell = ws.cell(current_row, col_idx)
        cell.fill = PatternFill(start_color=grade_colors[grade], end_color=grade_colors[grade], fill_type="solid")
        cell.border = thin_border
        col_idx += 1
        
        for subject in subjects:
            values = quarter_data[subject]
            subject_grades = []
            for val in values:
                if val:
                     parts = str(val).split(',')
                     for part in parts:
                         parsed_g = parse_grade(part.strip())
                         if parsed_g:
                             subject_grades.append(parsed_g)
            
            if grade == '5':
                count = sum(1 for g in subject_grades if g == 5)
            elif grade == '4':
                count = sum(1 for g in subject_grades if g == 4)
            else:
                count = sum(1 for g in subject_grades if g == 3)
            
            cell = ws.cell(current_row, col_idx)
            cell.value = count
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color=grade_colors[grade], end_color=grade_colors[grade], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            col_idx += 1
        
        # Пустые столбцы статистики
        for _ in range(3):
            cell = ws.cell(current_row, col_idx)
            cell.fill = PatternFill(start_color=grade_colors[grade], end_color=grade_colors[grade], fill_type="solid")
            cell.border = thin_border
            col_idx += 1
        
        current_row += 1
    
    # 5. Строки "Качество" и "Успеваемость"
    for row_name, bg_color in [('Качество', COLORS['quality_bg']), ('Успеваемость', COLORS['performance_bg'])]:
        col_idx = 1
        # Первая колонка - название строки
        cell = ws.cell(current_row, col_idx)
        cell.value = row_name
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col_idx += 1
        
        # Вторая колонка - пустая (для ФИО)
        cell = ws.cell(current_row, col_idx)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.border = thin_border
        col_idx += 1
        
        for subject in subjects:
            values = quarter_data[subject]
            subject_grades = []
            for val in values:
                if val:
                     parts = str(val).split(',')
                     for part in parts:
                         parsed_g = parse_grade(part.strip())
                         if parsed_g:
                             subject_grades.append(parsed_g)
            
            total = len(subject_grades)
            count_5 = sum(1 for g in subject_grades if g == 5)
            count_4 = sum(1 for g in subject_grades if g == 4)
            count_3 = sum(1 for g in subject_grades if g == 3)
            
            if row_name == 'Качество':
                value = round((count_5 + count_4) / total * 100, 2) if total > 0 else 0
            else:
                value = round((count_5 + count_4 + count_3) / total * 100, 2) if total > 0 else 0
            
            cell = ws.cell(current_row, col_idx)
            cell.value = value
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = '0.00'
            cell.border = thin_border
            col_idx += 1
        
        # Пустые столбцы статистики
        for _ in range(3):
            cell = ws.cell(current_row, col_idx)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.border = thin_border
            col_idx += 1
        
        current_row += 1
    
    # 6. Строки "Качество по классу" и "Успеваемость по классу"
    all_grades = []
    for subject in subjects:
        values = quarter_data[subject]
        for val in values:
            if val:
                 parts = str(val).split(',')
                 for part in parts:
                     parsed_g = parse_grade(part.strip())
                     if parsed_g:
                         all_grades.append(parsed_g)
    
    total_class = len(all_grades)
    count_5 = sum(1 for g in all_grades if g == 5)
    count_4 = sum(1 for g in all_grades if g == 4)
    count_3 = sum(1 for g in all_grades if g == 3)
    
    class_quality = round((count_5 + count_4) / total_class * 100, 2) if total_class > 0 else 0
    class_performance = round((count_5 + count_4 + count_3) / total_class * 100, 2) if total_class > 0 else 0
    
    for row_name, value in [('Качество по классу', class_quality), ('Успеваемость по классу', class_performance)]:
        col_idx = 1
        # Первая колонка - название строки
        cell = ws.cell(current_row, col_idx)
        cell.value = row_name
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=COLORS['class_stats_bg'], end_color=COLORS['class_stats_bg'], fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        
        # Вторая колонка - значение
        cell = ws.cell(current_row, col_idx + 1)
        cell.value = value
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=COLORS['class_stats_bg'], end_color=COLORS['class_stats_bg'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '0.00'
        cell.border = thin_border
        
        # Объединяем оставшиеся ячейки (начиная с 3-й колонки)
        if len(subjects) > 0:
            last_col = get_column_letter(stats_col_start + 2)
            ws.merge_cells(f'{get_column_letter(col_idx + 2)}{current_row}:{last_col}{current_row}')
            merged_cell = ws[f'{get_column_letter(col_idx + 2)}{current_row}']
            merged_cell.fill = PatternFill(start_color=COLORS['class_stats_bg'], end_color=COLORS['class_stats_bg'], fill_type="solid")
            merged_cell.border = thin_border
        
        current_row += 1
    
    # Отступ перед следующей четвертью
    return current_row + 2


def process_success_data(input_file='success_data.xlsx', output_file='processed_final.xlsx', class_name=None, output_dir=None):
    """Основная функция обработки данных"""
    print("="*70)
    print("ОБРАБОТКА ДАННЫХ ПО ЧЕТВЕРТЯМ")
    print("="*70)
    
    try:
        wb = load_workbook(input_file)
        print(f"\nЗагрузка файла: {input_file}")
        print(f"Найдено листов: {len(wb.sheetnames)}")
        
        output_wb = Workbook()
        output_wb.remove(output_wb.active)
        
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*70}")
            print(f"Обработка параллели: {sheet_name}")
            print(f"{'='*70}")
            
            try:
                # Читаем данные
                df, subjects_map = read_data_with_two_level_headers(input_file, sheet_name)
                print(f"Загружено записей: {len(df)}")
                print(f"Колонок: {len(df.columns)}")
                
                if len(df) == 0:
                    print("⚠ Нет данных для этой параллели")
                    continue
                
                # Находим столбец ФИО
                fio_column = None
                for col in df.columns:
                    if any(x in str(col).lower() for x in ['аты-жөні', 'фио', 'fio', 'аты']):
                        fio_column = col
                        break
                
                if fio_column is None and len(df.columns) > 1:
                    fio_column = df.columns[1]
                
                print(f"Столбец ФИО: {fio_column}")
                
                # Фильтруем строки с пустым ФИО
                if fio_column:
                    # Удаляем строки, где ФИО пустое или содержит только пробелы
                    before_filter = len(df)
                    df = df[df[fio_column].notna() & (df[fio_column].astype(str).str.strip() != '')].copy()
                    after_filter = len(df)
                    if before_filter != after_filter:
                        print(f"  Отфильтровано пустых строк: {before_filter - after_filter} (было {before_filter}, стало {after_filter})")
                
                if len(df) == 0:
                    print("⚠ Нет данных после фильтрации пустых строк")
                    continue
                
                # Обрезаем данные, если последовательность прервалась (проверяем по первой колонке с числами)
                # Ищем первую колонку, которая может содержать номера
                num_col = None
                for col in df.columns:
                    col_str = str(col).lower()
                    if 'параллель' in col_str or 'номер_строки' in col_str or 'аты-жөні' in col_str or 'фио' in col_str:
                        continue
                    # Берем первую колонку, которая может содержать числа
                    num_col = col
                    break
                
                if num_col is not None:
                    cut_row_idx = None
                    expected_num = 1
                    
                    for i, idx in enumerate(df.index):
                        value = df.loc[idx, num_col]
                        try:
                            if pd.isna(value) or str(value).strip() == '':
                                cut_row_idx = i
                                break
                            num_value = int(float(str(value).strip()))
                            if num_value != expected_num:
                                cut_row_idx = i
                                break
                            expected_num += 1
                        except (ValueError, TypeError):
                            cut_row_idx = i
                            break
                    
                    if cut_row_idx is not None:
                        original_len = len(df)
                        df = df.iloc[:cut_row_idx].copy()
                        print(f"  Данные обрезаны: {original_len} -> {len(df)} строк (последовательность прервалась)")
                
                # Извлекаем данные ФИО после фильтрации
                fio_data = df[fio_column].tolist() if fio_column else [''] * len(df)
                
                # Объединяем дублирующиеся столбцы
                print("Объединение дублирующихся столбцов...")
                print(f"  Всего колонок в DataFrame: {len(df.columns)}")
                print(f"  Первые 10 колонок: {list(df.columns[:10])}")
                
                # Проверяем наличие колонок с "Column" или "Колонка"
                column_cols = [col for col in df.columns if 'column' in str(col).lower() or 'колонка' in str(col).lower()]
                if column_cols:
                    print(f"  ⚠ Найдены колонки с 'Column'/'Колонка': {column_cols}")
                
                merged_data = merge_duplicate_columns(df)
                print(f"  Объединено предметов: {len(merged_data)}")
                print(f"  Предметы: {list(merged_data.keys())[:10]}")
                
                # Проверяем, не попали ли колонки с "Column" в merged_data
                column_subjects = [(s, q) for (s, q) in merged_data.keys() if s and ('column' in str(s).lower() or 'колонка' in str(s).lower())]
                if column_subjects:
                    print(f"  ⚠ ВНИМАНИЕ: В merged_data попали колонки с 'Column'/'Колонка': {column_subjects}")
                
                # Создаем лист
                clean_sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_')
                clean_sheet_name = clean_sheet_name.replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                if len(clean_sheet_name) > 31:
                    clean_sheet_name = clean_sheet_name[:31]
                
                # Проверяем, есть ли данные для хотя бы одной четверти
                has_data = False
                for quarter in QUARTERS_ORDER:
                    quarter_normalized = normalize_quarter(quarter)
                    # Проверяем, есть ли данные для этой четверти
                    for (subject, quarter_key), values in merged_data.items():
                        if quarter_key == quarter_normalized and any(v for v in values if v and str(v).strip()):
                            has_data = True
                            break
                    if has_data:
                        break
                
                if not has_data:
                    print(f"⚠ Нет данных для параллели {sheet_name}, пропускаем...")
                    continue
                
                output_ws = output_wb.create_sheet(title=clean_sheet_name)
                
                # Обрабатываем каждую четверть
                current_row = 1
                for quarter in QUARTERS_ORDER:
                    quarter_normalized = normalize_quarter(quarter)
                    print(f"  Обработка четверти: {quarter} (нормализовано: {quarter_normalized})")
                    current_row = create_quarter_table(output_ws, current_row, quarter_normalized, merged_data, fio_data)
                    print(f"    ✓ Данные для четверти {quarter_normalized} обработаны")
                
                # Проверяем, что лист не пустой
                if output_ws.max_row == 0 or output_ws.max_column == 0:
                    print(f"⚠ Лист для {sheet_name} оказался пустым, удаляем...")
                    output_wb.remove(output_ws)
                    continue
                
                # Настройка ширины колонок
                for col_idx in range(1, output_ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_idx == 1:
                        output_ws.column_dimensions[col_letter].width = 25
                    else:
                        max_length = 0
                        for row_idx in range(1, min(output_ws.max_row + 1, 100)):
                            cell_value = output_ws.cell(row_idx, col_idx).value
                            if cell_value:
                                max_length = max(max_length, len(str(cell_value)))
                        width = min(max_length + 2, 15) if max_length > 15 else min(max_length + 2, 12)
                        output_ws.column_dimensions[col_letter].width = width
                
                print(f"✓ Параллель {sheet_name} обработана")
                
            except Exception as e:
                print(f"✗ Ошибка при обработке параллели {sheet_name}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        # Проверяем, есть ли листы для сохранения
        if len(output_wb.sheetnames) == 0:
            print("\n⚠ Нет данных для сохранения (не создано ни одного листа)")
            print("Возможные причины:")
            print("  - Все параллели были пропущены из-за ошибок")
            print("  - Во всех параллелях отсутствуют данные")
            return False, None
        
        # Определяем имя выходного файла
        if class_name:
            # Очищаем имя класса от недопустимых символов для имени файла
            safe_class_name = re.sub(r'[<>:"/\\|?*]', '_', class_name)
            output_file = f"{safe_class_name}.xlsx"
        
        # Если указана папка для сохранения, добавляем её к пути
        if output_dir:
            import os
            output_file = os.path.join(output_dir, output_file)
        
        # Сохраняем файл
        print(f"\n{'='*70}")
        print(f"Сохранение файла: {output_file}")
        output_wb.save(output_file)
        print(f"✓ Файл успешно сохранен: {output_file}")
        print(f"  Создано листов: {len(output_wb.sheetnames)}")
        print("="*70)
        return True, output_file
        
    except Exception as e:
        print(f"✗ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False, None


if __name__ == '__main__':
    process_success_data(
        input_file='success_data.xlsx',
        output_file='processed_final.xlsx'
    )
