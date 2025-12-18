# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('example.xlsx')
ws = wb.active

output = []
output.append(f'Sheet: {ws.title}')
output.append(f'Max row: {ws.max_row}')
output.append(f'Max col: {ws.max_column}')
output.append('\n=== MERGED CELLS ===')
for merged in ws.merged_cells.ranges:
    output.append(f'Merged: {merged}')

output.append('\n=== FIRST ROW (HEADERS) ===')
for col in range(1, min(30, ws.max_column + 1)):
    cell = ws.cell(1, col)
    value = cell.value if cell.value else ''
    output.append(f'Col {col} ({get_column_letter(col)}): "{value}"')

output.append('\n=== SECOND ROW (HEADERS) ===')
for col in range(1, min(30, ws.max_column + 1)):
    cell = ws.cell(2, col)
    value = cell.value if cell.value else ''
    output.append(f'Col {col} ({get_column_letter(col)}): "{value}"')

output.append('\n=== FIRST 5 DATA ROWS ===')
for row in range(3, min(8, ws.max_row + 1)):
    output.append(f'\nRow {row}:')
    for col in range(1, min(30, ws.max_column + 1)):
        cell = ws.cell(row, col)
        value = cell.value if cell.value else ''
        output.append(f'  Col {col} ({get_column_letter(col)}): "{value}"')

with open('example_structure.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print('\n'.join(output))
wb.close()
