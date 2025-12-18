# -*- coding: utf-8 -*-
"""
Парсер для сайта mektep.edu.kz/_monitor/
Извлекает данные успеваемости из таблицы "Сапа" для всех классов параллели
"""
import time
import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()

class MektepScraper:
    def __init__(self):
        """Инициализация парсера"""
        self.base_url = "https://mektep.edu.kz/_monitor/"
        self.login_url = f"{self.base_url}index.php"
        self.driver = None
        self.wait = None
        self.data = {}  # {parallel: {class_name: table_data}}
        
    def setup_driver(self):
        """Настройка браузера Chrome"""
        chrome_options = Options()
        # Отключаем автоматизацию
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        
        # Проверяем, запущено ли в headless режиме (для Render и других серверов)
        is_headless = os.getenv('HEADLESS', 'false').lower() == 'true'
        if is_headless:
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
        else:
            chrome_options.add_argument("--start-maximized")
        
        # Дополнительные опции для стабильности
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-software-rasterizer")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, 30)
        # Устанавливаем таймаут загрузки страницы
        self.driver.set_page_load_timeout(30)
        print("✓ Браузер запущен")
    
    def open_page(self, url):
        """Открытие страницы с умным ожиданием полной загрузки"""
        try:
            print(f"Открытие страницы: {url}")
            self.driver.get(url)
            
            # Умное ожидание загрузки страницы
            print("Ожидание загрузки страницы...")
            
            # 1. Ждем, пока DOM полностью загрузится
            self.wait.until(
                lambda driver: driver.execute_script("return document.readyState") == "complete"
            )
            
            # 2. Ждем, пока jQuery (если используется) завершит все запросы
            try:
                self.wait.until(
                    lambda driver: driver.execute_script("return (typeof jQuery === 'undefined' || jQuery.active === 0)")
                )
            except TimeoutException:
                # Если jQuery не используется или не завершился, продолжаем
                pass
            
            # 3. Ждем появления основного контента (карточки или таблицы)
            try:
                # Для страницы школы ищем элементы, которые должны появиться
                if "id_mektep=" in url:
                    # Ждем появления заголовка "Отчет" или таблицы с данными
                    self.wait.until(
                        EC.any_of(
                            EC.presence_of_element_located((By.XPATH, "//h3[contains(text(), 'Отчет')]")),
                            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Успеваемость')]")),
                            EC.presence_of_element_located((By.XPATH, "//table")),
                            EC.presence_of_element_located((By.CSS_SELECTOR, ".card, .card-body"))
                        )
                    )
                else:
                    # Для других страниц ждем появления таблицы
                    self.wait.until(
                        EC.presence_of_element_located((By.TAG_NAME, "table"))
                    )
            except TimeoutException:
                # Если не нашли специфичные элементы, продолжаем
                print("⚠ Некоторые элементы не найдены, но продолжаем...")
            
            # 4. Дополнительная пауза для полной загрузки динамического контента
            time.sleep(2)
            
            # 5. Проверяем, что URL изменился на ожидаемый
            current_url = self.driver.current_url
            if url not in current_url and "id_mektep=" in url:
                # Проверяем, что мы на странице школы (содержит id_mektep)
                if "id_mektep=" not in current_url:
                    print(f"⚠ URL не соответствует ожидаемому. Текущий: {current_url}")
            
            print("✓ Страница загружена")
            return True
            
        except TimeoutException:
            print("⚠ Превышено время ожидания загрузки страницы")
            # Проверяем, может быть страница все же загрузилась
            try:
                if self.driver.execute_script("return document.readyState") == "complete":
                    print("⚠ DOM загружен, но некоторые элементы не появились. Продолжаем...")
                    return True
            except:
                pass
            return False
        except Exception as e:
            print(f"✗ Ошибка при открытии страницы: {e}")
            return False
    
    def login(self):
        """Ожидание ручной авторизации пользователя с автоматической проверкой"""
        try:
            # Открываем страницу авторизации
            if not self.open_page(self.login_url):
                return False
            
            print(f"\n{'='*60}")
            print("Страница авторизации открыта в браузере")
            print("Пожалуйста, введите данные для авторизации вручную")
            print("Ожидание авторизации...")
            print(f"{'='*60}\n")
            
            # Автоматически проверяем авторизацию каждые 2 секунды
            max_attempts = 300  # Максимум 10 минут ожидания (300 * 2 секунды)
            attempt = 0
            
            while attempt < max_attempts:
                # Быстрая проверка авторизации
                if self.check_authentication_quick():
                    print("✓ Авторизация успешна!")
                    return True
                
                attempt += 1
                if attempt % 15 == 0:  # Каждые 30 секунд выводим сообщение
                    print(f"Ожидание авторизации... (попытка {attempt}/{max_attempts})")
                
                time.sleep(2)  # Ждем 2 секунды перед следующей проверкой
            
            print("✗ Превышено время ожидания авторизации")
            return False
            
        except Exception as e:
            print(f"✗ Ошибка при авторизации: {e}")
            return False
    
    def check_authentication_quick(self):
        """Быстрая проверка успешности авторизации (без долгих ожиданий)"""
        try:
            current_url = self.driver.current_url
            
            # Проверяем, что мы не на странице логина
            if "index.php" in current_url and "login" not in current_url.lower():
                # Быстрая проверка URL
                if "pg_" in current_url or "reports" in current_url.lower():
                    return True
                
                # Быстрая проверка наличия элементов авторизованной страницы
                try:
                    # Используем короткий таймаут для быстрой проверки
                    quick_wait = WebDriverWait(self.driver, 1)
                    quick_wait.until(
                        EC.any_of(
                            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'pg_reports')]")),
                            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'pg_')]")),
                            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Отчеты')]"))
                        )
                    )
                    return True
                except (TimeoutException, NoSuchElementException):
                    # Если не нашли элементы, но URL изменился, считаем авторизацию успешной
                    if current_url != self.login_url and "index.php" in current_url:
                        return True
                    return False
            else:
                # Если все еще на странице логина, проверяем наличие сообщения об ошибке
                try:
                    error_message = self.driver.find_element(
                        By.CSS_SELECTOR, 
                        ".error, .alert-danger, [class*='error'], [class*='alert']"
                    )
                    if error_message and error_message.is_displayed():
                        return False
                except NoSuchElementException:
                    pass
                
                return False
                
        except Exception:
            return False
    
    def check_authentication(self):
        """Проверка успешности авторизации"""
        try:
            # Ждем изменения URL или появления элементов, указывающих на успешную авторизацию
            time.sleep(2)
            current_url = self.driver.current_url
            
            # Проверяем, что мы не на странице логина
            if "index.php" in current_url and "login" not in current_url.lower():
                # Проверяем наличие элементов, которые появляются после авторизации
                # (например, меню, ссылки на отчеты и т.д.)
                try:
                    # Ждем появления любого элемента, указывающего на авторизованную страницу
                    self.wait.until(
                        EC.any_of(
                            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'pg_reports')]")),
                            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'pg_')]")),
                            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Отчеты')]")),
                            EC.url_contains("pg_")
                        )
                    )
                    print("✓ Авторизация успешна")
                    return True
                except TimeoutException:
                    # Если не нашли специфичные элементы, проверяем URL
                    if "pg_" in current_url or "reports" in current_url.lower():
                        print("✓ Авторизация успешна (по URL)")
                        return True
                    else:
                        print("⚠ Не удалось подтвердить авторизацию, но продолжаем работу")
                        return True  # Продолжаем в любом случае
            else:
                # Если все еще на странице логина, проверяем наличие сообщения об ошибке
                try:
                    error_message = self.driver.find_element(
                        By.CSS_SELECTOR, 
                        ".error, .alert-danger, [class*='error'], [class*='alert']"
                    )
                    if error_message:
                        print(f"✗ Ошибка авторизации: {error_message.text}")
                        return False
                except NoSuchElementException:
                    pass
                
                print("⚠ Возможно, авторизация не прошла. Проверьте учетные данные.")
                return False
                
        except Exception as e:
            print(f"⚠ Ошибка при проверке авторизации: {e}")
            return False
    
    def navigate_to_reports(self):
        """Переход на страницу отчетов"""
        try:
            reports_url = f"{self.base_url}pg_reports.php"
            print(f"\nПереход на страницу отчетов: {reports_url}")
            
            if not self.open_page(reports_url):
                return False
            
            # Ждем загрузки таблицы со школами
            print("Ожидание загрузки таблицы школ...")
            try:
                # Ждем появления таблицы
                self.wait.until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                time.sleep(2)  # Дополнительная пауза для полной загрузки таблицы
                print("✓ Таблица школ загружена")
                return True
            except TimeoutException:
                print("⚠ Таблица не найдена, но продолжаем...")
                return True  # Продолжаем в любом случае
                
        except Exception as e:
            print(f"✗ Ошибка при переходе на страницу отчетов: {e}")
            return False
    
    def get_schools_list(self):
        """Получение списка школ из таблицы на странице pg_reports.php"""
        try:
            print("\nПоиск таблицы со списком школ...")
            
            # Ждем загрузки таблицы
            try:
                self.wait.until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                time.sleep(2)  # Дополнительная пауза для полной загрузки таблицы
            except TimeoutException:
                print("⚠ Таблица не найдена")
            
            # Ищем все таблицы на странице
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            
            if not tables:
                print("✗ Таблицы не найдены на странице")
                return []
            
            schools = []
            
            # Ищем таблицу со школами (обычно это первая большая таблица)
            for table in tables:
                try:
                    # Ищем заголовки таблицы
                    headers = table.find_elements(By.TAG_NAME, "th")
                    header_texts = [h.text.strip() for h in headers]
                    
                    # Проверяем, есть ли столбец "Районы/города/школы" или похожий
                    if any("Районы" in text or "города" in text or "школы" in text or "Школа" in text 
                           for text in header_texts):
                        print(f"✓ Найдена таблица со школами (столбцов: {len(headers)})")
                        
                        # Находим индекс столбца со школами
                        school_col_index = None
                        for idx, header_text in enumerate(header_texts):
                            if "Районы" in header_text or "города" in header_text or "школы" in header_text or "Школа" in header_text:
                                school_col_index = idx
                                break
                        
                        if school_col_index is None:
                            # Если не нашли по тексту, берем первый столбец
                            school_col_index = 0
                        
                        # Получаем все строки таблицы (кроме заголовка)
                        rows = table.find_elements(By.TAG_NAME, "tr")
                        
                        for row_idx, row in enumerate(rows[1:], start=1):  # Пропускаем заголовок
                            try:
                                cells = row.find_elements(By.TAG_NAME, "td")
                                if len(cells) > school_col_index:
                                    school_cell = cells[school_col_index]
                                    
                                    # Ищем ссылку в ячейке
                                    try:
                                        link = school_cell.find_element(By.TAG_NAME, "a")
                                        school_name = link.text.strip()
                                        school_url = link.get_attribute("href")
                                        
                                        if school_name and school_url:
                                            # Нормализуем URL (может быть относительным)
                                            if not school_url.startswith("http"):
                                                if school_url.startswith("/"):
                                                    school_url = f"https://mektep.edu.kz{school_url}"
                                                else:
                                                    school_url = f"{self.base_url}{school_url}"
                                            
                                            schools.append({
                                                "index": len(schools) + 1,
                                                "name": school_name,
                                                "url": school_url
                                            })
                                    except NoSuchElementException:
                                        continue
                            except (NoSuchElementException, IndexError):
                                continue
                        
                        break  # Нашли нужную таблицу, выходим
                        
                except Exception as e:
                    continue
            
            if not schools:
                # Если не нашли по заголовкам, пробуем найти все ссылки в таблицах
                print("Поиск школ альтернативным методом...")
                for table in tables:
                    try:
                        links = table.find_elements(By.TAG_NAME, "a")
                        for link in links:
                            href = link.get_attribute("href")
                            text = link.text.strip()
                            # Проверяем, что это ссылка на школу (содержит id_mektep или pg_reports)
                            if href and ("id_mektep=" in href or "pg_reports" in href) and text:
                                # Нормализуем URL
                                if not href.startswith("http"):
                                    if href.startswith("/"):
                                        href = f"https://mektep.edu.kz{href}"
                                    else:
                                        href = f"{self.base_url}{href}"
                                
                                # Проверяем, что такой школы еще нет в списке
                                if not any(s["url"] == href for s in schools):
                                    schools.append({
                                        "index": len(schools) + 1,
                                        "name": text,
                                        "url": href
                                    })
                    except Exception:
                        continue
            
            print(f"✓ Найдено школ: {len(schools)}")
            return schools
            
        except Exception as e:
            print(f"✗ Ошибка при получении списка школ: {e}")
            return []
    
    def select_school(self, school_index=None):
        """Выбор школы из списка и переход по ссылке"""
        try:
            # Получаем список школ
            schools = self.get_schools_list()
            
            if not schools:
                print("✗ Список школ пуст")
                return False
            
            # Выводим список школ
            print(f"\n{'='*60}")
            print("СПИСОК ШКОЛ:")
            print(f"{'='*60}")
            for school in schools:
                print(f"{school['index']}. {school['name']}")
            print(f"{'='*60}\n")
            
            # Выбор школы
            if school_index is None:
                # Интерактивный выбор
                while True:
                    try:
                        choice = input(f"Выберите номер школы (1-{len(schools)}): ").strip()
                        school_index = int(choice)
                        if 1 <= school_index <= len(schools):
                            break
                        else:
                            print(f"Пожалуйста, введите число от 1 до {len(schools)}")
                    except ValueError:
                        print("Пожалуйста, введите корректный номер")
                    except KeyboardInterrupt:
                        print("\nОтменено пользователем")
                        return False
            
            # Проверяем корректность индекса
            if not (1 <= school_index <= len(schools)):
                print(f"✗ Неверный номер школы. Допустимый диапазон: 1-{len(schools)}")
                return False
            
            # Получаем выбранную школу
            selected_school = schools[school_index - 1]
            print(f"\nВыбрана школа: {selected_school['name']}")
            print(f"Переход по ссылке: {selected_school['url']}")
            
            # Переходим напрямую по URL (надежнее, чем клик по элементу)
            if not self.open_page(selected_school['url']):
                return False
            
            # Дополнительная проверка загрузки страницы школы
            print("Проверка загрузки контента страницы школы...")
            try:
                # Ждем появления ключевых элементов страницы школы
                self.wait.until(
                    EC.any_of(
                        EC.presence_of_element_located((By.XPATH, "//h3[contains(text(), 'Отчет')]")),
                        EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Успеваемость')]")),
                        EC.presence_of_element_located((By.XPATH, "//div[@id='sapa_view']")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".card-body"))
                    )
                )
                print("✓ Контент страницы школы загружен")
            except TimeoutException:
                print("⚠ Некоторые элементы не найдены, но продолжаем...")
            
            print(f"✓ Переход на страницу школы выполнен")
            return True
            
        except Exception as e:
            print(f"✗ Ошибка при выборе школы: {e}")
            return False
    
    def get_classes_list(self):
        """Получение списка классов из навигационных вкладок (pills)"""
        try:
            print("\nПоиск списка классов...")
            
            # Ждем появления навигационных вкладок с классами
            try:
                # Ищем элемент с id="pills-tab" или классом "nav nav-pills"
                self.wait.until(
                    EC.any_of(
                        EC.presence_of_element_located((By.ID, "pills-tab")),
                        EC.presence_of_element_located((By.CSS_SELECTOR, "ul.nav.nav-pills"))
                    )
                )
                time.sleep(1)  # Дополнительная пауза
            except TimeoutException:
                print("⚠ Навигационные вкладки не найдены")
            
            classes = []
            
            # Ищем список классов (ul с id="pills-tab" или классом "nav nav-pills")
            try:
                # Сначала пробуем найти по ID
                pills_tab = None
                try:
                    pills_tab = self.driver.find_element(By.ID, "pills-tab")
                except NoSuchElementException:
                    # Если не нашли по ID, ищем по классу
                    pills_tab = self.driver.find_element(By.CSS_SELECTOR, "ul.nav.nav-pills")
                
                if pills_tab:
                    # Находим все ссылки внутри списка
                    links = pills_tab.find_elements(By.TAG_NAME, "a")
                    
                    for link in links:
                        try:
                            class_text = link.text.strip()
                            # Извлекаем номер класса из текста (например, "11  класс" -> "11")
                            # Ищем число в начале текста
                            match = re.match(r'(\d+)', class_text)
                            if match:
                                class_number = match.group(1)
                                
                                # Получаем href (может быть якорем или относительным URL)
                                href = link.get_attribute("href")
                                
                                # Проверяем, активна ли вкладка
                                is_active = "active" in link.get_attribute("class")
                                
                                classes.append({
                                    "index": len(classes) + 1,
                                    "number": class_number,
                                    "name": class_text,
                                    "href": href,
                                    "is_active": is_active,
                                    "element": link  # Сохраняем элемент для клика
                                })
                        except Exception as e:
                            continue
                    
                    # Сортируем классы по номеру (от большего к меньшему: 11, 10, 9...)
                    classes.sort(key=lambda x: int(x["number"]), reverse=True)
                    
                    # Обновляем индексы после сортировки
                    for idx, cls in enumerate(classes, start=1):
                        cls["index"] = idx
                    
                    print(f"✓ Найдено классов: {len(classes)}")
                    return classes
                else:
                    print("✗ Список классов не найден")
                    return []
                    
            except NoSuchElementException:
                print("✗ Элемент со списком классов не найден")
                return []
            
        except Exception as e:
            print(f"✗ Ошибка при получении списка классов: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def select_class_tab(self, class_number):
        """Выбор вкладки класса (клик по вкладке с номером класса)"""
        try:
            print(f"\nВыбор вкладки класса: {class_number}")
            
            # Ищем вкладку с нужным номером класса
            # Ищем ссылку, которая содержит номер класса
            link_xpath = f"//ul[@id='pills-tab']//a[contains(text(), '{class_number}')] | //ul[contains(@class, 'nav-pills')]//a[contains(text(), '{class_number}')]"
            
            try:
                class_link = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, link_xpath))
                )
                
                # Прокручиваем к элементу
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", class_link)
                time.sleep(0.5)
                
                # Используем JavaScript для клика, чтобы обойти проблему с перекрытием
                self.driver.execute_script("arguments[0].click();", class_link)
                print(f"✓ Клик по вкладке '{class_number} класс' выполнен (через JavaScript)")
                
                # Ждем загрузки таблицы с классами
                time.sleep(2)
                
                # Ждем появления таблицы с классами
                try:
                    self.wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped, table.table-bordered"))
                    )
                    print("✓ Таблица с классами загружена")
                except TimeoutException:
                    print("⚠ Таблица не найдена, но продолжаем...")
                
                return True
                
            except TimeoutException:
                print(f"✗ Вкладка класса '{class_number}' не найдена")
                return False
                
        except Exception as e:
            print(f"✗ Ошибка при выборе вкладки класса: {e}")
            return False
    
    def get_class_groups_from_table(self):
        """Получение списка классов (групп) из таблицы после выбора вкладки класса"""
        try:
            print("\nПоиск таблицы с классами...")
            
            # Ждем появления таблицы
            try:
                self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped, table.table-bordered"))
                )
                time.sleep(1)
            except TimeoutException:
                print("⚠ Таблица не найдена")
            
            class_groups = []
            
            # Ищем таблицу с классами
            tables = self.driver.find_elements(By.CSS_SELECTOR, "table.table-striped, table.table-bordered")
            
            for table in tables:
                try:
                    # Ищем заголовки таблицы - сначала в thead, потом в первой строке
                    headers = []
                    thead = table.find_elements(By.TAG_NAME, "thead")
                    if thead:
                        # Заголовки в thead
                        header_row = thead[0].find_elements(By.TAG_NAME, "tr")
                        if header_row:
                            headers = header_row[0].find_elements(By.TAG_NAME, "td")
                            if not headers:
                                headers = header_row[0].find_elements(By.TAG_NAME, "th")
                    
                    # Если не нашли в thead, ищем в первой строке tbody или просто первой строке
                    if not headers:
                        rows = table.find_elements(By.TAG_NAME, "tr")
                        if rows:
                            headers = rows[0].find_elements(By.TAG_NAME, "td")
                            if not headers:
                                headers = rows[0].find_elements(By.TAG_NAME, "th")
                    
                    if not headers:
                        continue
                    
                    header_texts = [h.text.strip() for h in headers]
                    
                    # Проверяем, что это таблица с классами (есть столбец "Класс")
                    if "Класс" in header_texts or "класс" in " ".join(header_texts).lower():
                        print(f"✓ Найдена таблица с классами (столбцов: {len(headers)})")
                        print(f"  Заголовки: {header_texts}")
                        
                        # Находим индексы столбцов
                        class_col_idx = None
                        type_col_idx = None
                        language_col_idx = None
                        shift_col_idx = None
                        teacher_col_idx = None
                        students_col_idx = None
                        actions_col_idx = None
                        
                        for idx, header_text in enumerate(header_texts):
                            header_lower = header_text.lower()
                            if "класс" in header_lower and "тип" not in header_lower and class_col_idx is None:
                                class_col_idx = idx
                            elif "тип" in header_lower and "класс" in header_lower:
                                type_col_idx = idx
                            elif "язык" in header_lower:
                                language_col_idx = idx
                            elif "смена" in header_lower:
                                shift_col_idx = idx
                            elif "руководитель" in header_lower or "классный" in header_lower:
                                teacher_col_idx = idx
                            elif "учащиеся" in header_lower:
                                students_col_idx = idx
                            elif "действия" in header_lower:
                                actions_col_idx = idx
                        
                        # Если не нашли индекс класса, используем первый столбец
                        if class_col_idx is None:
                            class_col_idx = 0
                        
                        print(f"  Индексы столбцов: Класс={class_col_idx}, Тип={type_col_idx}, Язык={language_col_idx}, "
                              f"Смена={shift_col_idx}, Руководитель={teacher_col_idx}, Учащиеся={students_col_idx}, Действия={actions_col_idx}")
                        
                        # Получаем все строки таблицы
                        rows = table.find_elements(By.TAG_NAME, "tr")
                        
                        # Определяем, с какой строки начинаются данные (пропускаем заголовок)
                        start_row = 1 if len(rows) > 1 else 0
                        
                        for row_idx, row in enumerate(rows[start_row:], start=start_row):
                            try:
                                cells = row.find_elements(By.TAG_NAME, "td")
                                
                                if len(cells) == 0:
                                    continue
                                
                                # Извлекаем данные из ячеек
                                class_name = ""
                                if class_col_idx is not None and len(cells) > class_col_idx:
                                    class_name = cells[class_col_idx].text.strip()
                                
                                class_type = ""
                                if type_col_idx is not None and len(cells) > type_col_idx:
                                    class_type = cells[type_col_idx].text.strip()
                                
                                language = ""
                                if language_col_idx is not None and len(cells) > language_col_idx:
                                    language = cells[language_col_idx].text.strip()
                                
                                shift = ""
                                if shift_col_idx is not None and len(cells) > shift_col_idx:
                                    shift = cells[shift_col_idx].text.strip()
                                
                                teacher = ""
                                if teacher_col_idx is not None and len(cells) > teacher_col_idx:
                                    teacher = cells[teacher_col_idx].text.strip()
                                
                                students = ""
                                if students_col_idx is not None and len(cells) > students_col_idx:
                                    students = cells[students_col_idx].text.strip()
                                
                                # Извлекаем литера класса из названия (например, "11 «А»" -> "А")
                                class_letter = ""
                                if class_name:
                                    letter_match = re.search(r'[«"]?([А-ЯЁA-Z])[«"]?', class_name)
                                    if letter_match:
                                        class_letter = letter_match.group(1)
                                
                                # Ищем кнопку в столбце "Действия"
                                button = None
                                if actions_col_idx is not None and len(cells) > actions_col_idx:
                                    try:
                                        button = cells[actions_col_idx].find_element(By.TAG_NAME, "button")
                                    except NoSuchElementException:
                                        pass
                                
                                # Добавляем класс, если есть название
                                if class_name:
                                    class_groups.append({
                                        "index": len(class_groups) + 1,
                                        "name": class_name,
                                        "letter": class_letter,
                                        "type": class_type,
                                        "language": language,
                                        "shift": shift,
                                        "teacher": teacher,
                                        "students": students,
                                        "button": button
                                    })
                                    print(f"  Найден класс: {class_name} (Литера: {class_letter})")
                                else:
                                    # Отладочная информация
                                    if row_idx < 5:  # Выводим только для первых строк
                                        cell_texts = [cell.text.strip() for cell in cells]
                                        print(f"  Строка {row_idx}: {cell_texts} (нет названия класса)")
                                        
                            except Exception as e:
                                if row_idx < 5:  # Выводим только для первых ошибок
                                    print(f"  Ошибка в строке {row_idx}: {e}")
                                continue
                        
                        break  # Нашли нужную таблицу
                        
                except Exception as e:
                    continue
            
            if class_groups:
                print(f"✓ Найдено классов в таблице: {len(class_groups)}")
            else:
                print("✗ Классы в таблице не найдены")
            
            return class_groups
            
        except Exception as e:
            print(f"✗ Ошибка при получении классов из таблицы: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def select_class_group(self, class_groups, class_group_index=None):
        """Выбор конкретного класса из списка и клик по кнопке 'Успеваемость'"""
        try:
            if not class_groups:
                print("✗ Список классов пуст")
                return False
            
            # Выводим список классов
            print(f"\n{'='*60}")
            print("ВЫБЕРИТЕ КОНКРЕТНЫЙ КЛАСС:")
            print(f"{'='*60}")
            for group in class_groups:
                print(f"{group['index']}. {group['name']} (Литера: {group['letter']}, "
                      f"Учащиеся: {group['students']}, Руководитель: {group['teacher']})")
            print(f"{'='*60}\n")
            
            # Выбор класса
            if class_group_index is None:
                while True:
                    try:
                        choice = input(f"Выберите номер класса (1-{len(class_groups)}): ").strip()
                        class_group_index = int(choice)
                        if 1 <= class_group_index <= len(class_groups):
                            break
                        else:
                            print(f"Пожалуйста, введите число от 1 до {len(class_groups)}")
                    except ValueError:
                        print("Пожалуйста, введите корректный номер")
                    except KeyboardInterrupt:
                        print("\nОтменено пользователем")
                        return False
            
            # Проверяем корректность индекса
            if not (1 <= class_group_index <= len(class_groups)):
                print(f"✗ Неверный номер класса. Допустимый диапазон: 1-{len(class_groups)}")
                return False
            
            # Получаем выбранный класс
            selected_group = class_groups[class_group_index - 1]
            print(f"\nВыбран класс: {selected_group['name']}")
            
            # Кликаем по кнопке "Успеваемость"
            if selected_group['button']:
                try:
                    # Прокручиваем к кнопке
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selected_group['button'])
                    time.sleep(0.5)
                    
                    # Используем JavaScript для клика, чтобы обойти проблемы с перекрытием
                    self.driver.execute_script("arguments[0].click();", selected_group['button'])
                    print("✓ Клик по кнопке 'Успеваемость' выполнен")
                    
                    # Ждем открытия модального окна
                    time.sleep(1)
                    
                    # Ждем появления и открытия модального окна
                    try:
                        # Ждем, пока модальное окно станет видимым (Bootstrap добавляет класс "show")
                        self.wait.until(
                            EC.any_of(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa.modal.show")),
                                EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa.modal:not(.fade)")),
                                EC.presence_of_element_located((By.ID, "classSapa"))
                            )
                        )
                        
                        # Дополнительно ждем загрузки содержимого (таблица может загружаться через AJAX)
                        time.sleep(2)
                        
                        # Ждем появления таблицы внутри модального окна
                        try:
                            self.wait.until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa table"))
                            )
                            print("✓ Модальное окно 'Сапа' открыто и таблица загружена")
                        except TimeoutException:
                            print("⚠ Таблица не появилась в модальном окне, но продолжаем...")
                        
                        return selected_group  # Возвращаем информацию о выбранном классе
                    except TimeoutException:
                        print("⚠ Модальное окно не найдено, но продолжаем...")
                        return selected_group
                        
                except Exception as e:
                    print(f"✗ Ошибка при клике по кнопке: {e}")
                    return False
            else:
                print("✗ Кнопка 'Успеваемость' не найдена для выбранного класса")
                return False
                
        except Exception as e:
            print(f"✗ Ошибка при выборе класса: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def is_modal_open(self):
        """Проверка, открыто ли модальное окно"""
        try:
            modal = self.driver.find_element(By.ID, "classSapa")
            # Проверяем наличие класса "show" или отсутствие класса "fade"
            classes = modal.get_attribute("class") or ""
            if "show" in classes or ("modal" in classes and "fade" not in classes):
                # Также проверяем наличие backdrop
                try:
                    backdrop = self.driver.find_element(By.CSS_SELECTOR, ".modal-backdrop")
                    return True
                except NoSuchElementException:
                    # Если нет backdrop, но есть класс show - все равно считаем открытым
                    return "show" in classes
            return False
        except NoSuchElementException:
            return False
        except Exception as e:
            print(f"⚠ Ошибка при проверке модального окна: {e}")
            return False
    
    def is_modal_closed(self):
        """Проверка, закрыто ли модальное окно"""
        try:
            # Проверяем отсутствие класса "show" у модального окна
            modal = self.driver.find_element(By.ID, "classSapa")
            classes = modal.get_attribute("class") or ""
            has_show = "show" in classes
            
            # Проверяем отсутствие backdrop
            try:
                backdrop = self.driver.find_element(By.CSS_SELECTOR, ".modal-backdrop")
                has_backdrop = True
            except NoSuchElementException:
                has_backdrop = False
            
            # Модальное окно закрыто, если нет класса "show" и нет backdrop
            return not has_show and not has_backdrop
        except NoSuchElementException:
            # Если модальное окно вообще не найдено, считаем его закрытым
            return True
        except Exception as e:
            print(f"⚠ Ошибка при проверке закрытия модального окна: {e}")
            return False
    
    def close_modal(self):
        """Закрытие модального окна с проверкой"""
        try:
            # Пробуем найти и кликнуть по кнопке закрытия
            try:
                close_button = self.driver.find_element(By.CSS_SELECTOR, "#classSapa .modal-header .close, #classSapa button[data-dismiss='modal'], #classSapa .close")
                if close_button:
                    # Используем JavaScript для клика
                    self.driver.execute_script("arguments[0].click();", close_button)
                    time.sleep(0.2)  # Уменьшена пауза
                    # Проверяем, что закрылось
                    if self.is_modal_closed():
                        print("✓ Модальное окно закрыто")
                        return True
            except NoSuchElementException:
                pass
            
            # Если кнопка не найдена, пробуем закрыть через клик вне модального окна
            try:
                # Кликаем по backdrop (затемненной области)
                backdrop = self.driver.find_element(By.CSS_SELECTOR, ".modal-backdrop")
                if backdrop:
                    self.driver.execute_script("arguments[0].click();", backdrop)
                    time.sleep(0.2)  # Уменьшена пауза
                    if self.is_modal_closed():
                        print("✓ Модальное окно закрыто (клик по backdrop)")
                        return True
            except NoSuchElementException:
                pass
            
            # Последняя попытка - через JavaScript закрыть модальное окно
            try:
                self.driver.execute_script("$('#classSapa').modal('hide');")
                time.sleep(0.3)  # Уменьшена пауза
                # Ждем, пока модальное окно закроется (с таймаутом)
                wait = WebDriverWait(self.driver, 3)  # Уменьшен таймаут
                wait.until(lambda d: self.is_modal_closed())
                print("✓ Модальное окно закрыто (JavaScript)")
                return True
            except TimeoutException:
                print("⚠ Модальное окно не закрылось через JavaScript, пробуем принудительно...")
                # Принудительное закрытие через удаление классов
                try:
                    self.driver.execute_script("""
                        var modal = document.getElementById('classSapa');
                        if (modal) {
                            modal.classList.remove('show');
                            modal.style.display = 'none';
                        }
                        var backdrop = document.querySelector('.modal-backdrop');
                        if (backdrop) {
                            backdrop.remove();
                        }
                        document.body.classList.remove('modal-open');
                    """)
                    time.sleep(0.2)  # Уменьшена пауза
                    if self.is_modal_closed():
                        print("✓ Модальное окно закрыто (принудительно)")
                        return True
                except:
                    pass
            except:
                pass
                
        except Exception as e:
            print(f"⚠ Ошибка при закрытии модального окна: {e}")
        
        return False
    
    def extract_modal_table_data(self):
        """Извлечение данных из таблицы 'Сапа' в модальном окне"""
        try:
            print("\nИзвлечение данных из таблицы 'Сапа' в модальном окне...")
            
            # Ждем появления модального окна и его открытия
            try:
                modal = self.wait.until(
                    EC.presence_of_element_located((By.ID, "classSapa"))
                )
                
                # Проверяем, что модальное окно видимо (открыто)
                # Модальное окно Bootstrap имеет классы "modal fade show" когда открыто
                # Уменьшаем паузу - используем ожидание вместо фиксированной задержки
                
                # Ждем появления таблицы внутри модального окна
                # Таблица может загружаться динамически через AJAX
                try:
                    self.wait.until(
                        EC.any_of(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa table")),
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa .modal-body table")),
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa .modal-content table"))
                        )
                    )
                    # Ждем, пока таблица полностью загрузится (проверяем наличие строк)
                    self.wait.until(lambda d: len(d.find_elements(By.CSS_SELECTOR, "#classSapa table tbody tr")) > 0)
                    time.sleep(0.3)  # Минимальная пауза для стабильности
                except TimeoutException:
                    print("⚠ Таблица не появилась в модальном окне, ищем альтернативным способом...")
                    time.sleep(0.5)  # Небольшая пауза перед альтернативным поиском
                
            except TimeoutException:
                print("✗ Модальное окно не найдено")
                return None
            
            # Ищем таблицу внутри модального окна
            table = None
            try:
                # Ищем таблицу внутри модального окна #classSapa
                modal = self.driver.find_element(By.ID, "classSapa")
                
                # Пробуем найти таблицу в modal-body или modal-content
                try:
                    table = modal.find_element(By.CSS_SELECTOR, ".modal-body table, .modal-content table")
                except NoSuchElementException:
                    # Ищем любую таблицу внутри модального окна
                    table = modal.find_element(By.TAG_NAME, "table")
                    
            except NoSuchElementException:
                print("✗ Таблица не найдена в модальном окне")
                # Пробуем найти таблицу на странице (возможно, она не в модальном окне)
                try:
                    table = self.driver.find_element(By.CSS_SELECTOR, "table.table-hover.table-responsive.table-bordered")
                except NoSuchElementException:
                    pass
            
            if not table:
                print("✗ Таблица 'Сапа' не найдена")
                return None
            
            print("✓ Таблица 'Сапа' найдена в модальном окне")
            
            # Используем оптимизированный метод извлечения через JavaScript
            # Это намного быстрее, чем итерация через Selenium
            try:
                headers_data, table_data = self._extract_table_data_fast(table)
                print("✓ Данные извлечены через оптимизированный метод")
            except Exception as e:
                print(f"⚠ Ошибка при быстром извлечении, используем стандартный метод: {e}")
                # Fallback на старый метод
                headers_data = self._extract_table_headers(table)
                table_data = self._extract_table_body(table)
            
            # Проверяем соответствие количества колонок
            if table_data and headers_data:
                expected_cols = 2 + len(headers_data.get("second_row", []))  # 2 служебных + предметы * 5
                if table_data:
                    actual_cols = len(table_data[0])
                    if actual_cols != expected_cols:
                        print(f"  - ⚠ ВНИМАНИЕ: Несоответствие колонок!")
                        print(f"    Ожидается: {expected_cols} (2 служебных + {len(headers_data.get('second_row', []))} колонок по предметам)")
                        print(f"    Получено: {actual_cols} колонок в данных")
                        print(f"    Разница: {actual_cols - expected_cols} колонок")
                    else:
                        print(f"  - ✓ Количество колонок соответствует: {expected_cols}")
            
            return {
                "headers": headers_data,
                "data": table_data
            }
            
        except TimeoutException:
            print("✗ Таблица не найдена")
            return None
        except Exception as e:
            print(f"✗ Ошибка при извлечении данных: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_table_data_fast(self, table):
        """Быстрое извлечение данных таблицы через JavaScript (оптимизированный метод)"""
        try:
            # Используем JavaScript для извлечения всей таблицы сразу
            table_html = self.driver.execute_script("""
                var table = arguments[0];
                if (!table) return null;
                
                // Извлекаем заголовки
                var thead = table.querySelector('thead');
                var headers = {
                    first_row: [],
                    second_row: [],
                    subjects: [],
                    first_col_name: "№",
                    second_col_name: "Аты-жөні"
                };
                
                if (thead) {
                    var rows = thead.querySelectorAll('tr');
                    if (rows.length >= 2) {
                        // Первая строка - предметы
                        var firstRow = rows[0];
                        var firstCells = firstRow.querySelectorAll('th');
                        var cellIndex = 0;
                        
                        for (var i = 0; i < firstCells.length; i++) {
                            var cell = firstCells[i];
                            var text = cell.textContent.trim();
                            var colspan = parseInt(cell.getAttribute('colspan') || '1');
                            var rowspan = parseInt(cell.getAttribute('rowspan') || '1');
                            
                            if (cellIndex == 0) {
                                // Пропускаем первую пустую ячейку
                            } else if (cellIndex == 1 && rowspan > 1) {
                                headers.second_col_name = text || "Аты-жөні";
                            } else {
                                headers.first_row.push({
                                    text: text || "",
                                    colspan: colspan
                                });
                                for (var j = 0; j < colspan; j++) {
                                    headers.subjects.push(text || "");
                                }
                            }
                            cellIndex++;
                        }
                        
                        // Вторая строка - четверти
                        var secondRow = rows[1];
                        var secondCells = secondRow.querySelectorAll('th');
                        var rawQuarters = [];
                        for (var i = 0; i < secondCells.length; i++) {
                            var text = secondCells[i].textContent.trim();
                            if (text) rawQuarters.push(text);
                        }
                        
                        // Определяем уникальные четверти
                        var baseQuarters = [];
                        for (var i = 0; i < rawQuarters.length; i++) {
                            if (baseQuarters.indexOf(rawQuarters[i]) === -1) {
                                baseQuarters.push(rawQuarters[i]);
                            }
                        }
                        if (baseQuarters.length === 0) {
                            baseQuarters = ["І", "ІІ", "ІІІ", "ІV", "Ж"];
                        }
                        
                        // Строим вторую строку заголовков
                        for (var i = 0; i < headers.first_row.length; i++) {
                            var colspan = headers.first_row[i].colspan || 1;
                            for (var j = 0; j < baseQuarters.length; j++) {
                                headers.second_row.push(baseQuarters[j]);
                            }
                        }
                        
                        headers.unique_quarters = baseQuarters;
                        headers.quarters_per_subject = baseQuarters.length;
                    }
                }
                
                // Извлекаем данные из tbody
                var tbody = table.querySelector('tbody');
                var tableData = [];
                
                if (tbody) {
                    var rows = tbody.querySelectorAll('tr');
                    for (var i = 0; i < rows.length; i++) {
                        var row = rows[i];
                        var cells = row.querySelectorAll('td');
                        
                        if (cells.length < 2) continue;
                        
                        // Пропускаем служебные строки
                        var firstCell = cells[0];
                        var colspan = firstCell.getAttribute('colspan');
                        if (colspan && parseInt(colspan) >= 999) continue;
                        
                        var rowClass = row.getAttribute('class') || '';
                        if (rowClass.indexOf('badge-') !== -1) {
                            var firstText = firstCell.textContent.trim();
                            if (['5', '4', '3', '2'].indexOf(firstText) !== -1) {
                                if (firstCell.querySelector('b')) continue;
                            }
                        }
                        
                        var firstText = firstCell.textContent.trim();
                        if (firstText === 'үлгерімі' || firstText === 'сапасы') continue;
                        
                        // Извлекаем данные строки
                        var rowData = [];
                        for (var j = 0; j < cells.length; j++) {
                            var text = cells[j].textContent.trim();
                            // Очищаем от лишних пробелов и переносов строк
                            text = text.replace(/\\s+/g, ' ');
                            rowData.push(text);
                        }
                        
                        // Проверяем, что это строка с данными ученика
                        if (rowData.length >= 2) {
                            var numValue = rowData[0];
                            var fioValue = rowData[1];
                            if (fioValue && fioValue.trim()) {
                                // Проверяем, что первая колонка - число или есть ФИО
                                if (!isNaN(numValue) || fioValue.trim()) {
                                    tableData.push(rowData);
                                }
                            }
                        }
                    }
                }
                
                return {
                    headers: headers,
                    data: tableData
                };
            """, table)
            
            if not table_html:
                raise Exception("Не удалось извлечь данные через JavaScript")
            
            headers_data = table_html.get("headers", {})
            table_data = table_html.get("data", [])
            
            # Отладочная информация
            print(f"✓ Извлечено через JavaScript:")
            print(f"  - Предметов: {len(headers_data.get('first_row', []))}")
            print(f"  - Четвертей: {len(headers_data.get('second_row', []))}")
            print(f"  - Строк данных: {len(table_data)}")
            
            return headers_data, table_data
            
        except Exception as e:
            print(f"⚠ Ошибка при быстром извлечении: {e}")
            raise
    
    def _extract_table_headers(self, table):
        """Извлечение двухуровневых заголовков таблицы 'Сапа' из модального окна.
        
        Структура в HTML (из сапа.html):
        - thead, первая строка:
          - <th rowspan="2"></th> (пустая первая ячейка)
          - <th rowspan="2">Аты-жөні</th> (вторая ячейка)
          - <th colspan="5">Предмет</th> (предметы с colspan=5)
        - thead, вторая строка:
          - (первые 2 ячейки пропущены из-за rowspan)
          - <th>І</th>, <th>ІІ</th>, <th>ІІІ</th>, <th>ІV</th>, <th>Ж</th> (четверти для каждого предмета)
        
        Требуемая структура Excel:
        - 1-й столбец: «№» (занимает 2 строки заголовка)
        - 2-й столбец: «Аты-жөні» (занимает 2 строки заголовка)
        - Начиная с 3-го столбца: предметы, каждый предмет занимает 5 столбцов (І, ІІ, ІІІ, ІV, Ж)
        """
        try:
            headers_data = {
                "first_row": [],      # Предметы с colspan
                "second_row": [],     # Четверти (І, ІІ, ІІІ, ІV, Ж) развёрнутые по всем предметам
                "subjects": [],       # Список предметов (по столбцам)
                "first_col_name": "№",        # Явно задаём первый столбец
                "second_col_name": "Аты-жөні" # Явно задаём второй столбец
            }
            
            # Ищем thead
            thead = table.find_elements(By.TAG_NAME, "thead")
            if not thead:
                return headers_data
            
            rows = thead[0].find_elements(By.TAG_NAME, "tr")
            if len(rows) < 2:
                return headers_data
            
            # --- Первая строка заголовков: предметы ---
            first_row = rows[0]
            first_cells = first_row.find_elements(By.TAG_NAME, "th")
            
            print(f"  - Ячеек в первой строке заголовков: {len(first_cells)}")
            
            cell_index = 0
            for cell in first_cells:
                text = cell.text.strip()
                colspan = int(cell.get_attribute("colspan") or "1")
                rowspan = int(cell.get_attribute("rowspan") or "1")
                
                if cell_index == 0:
                    # Первая ячейка - пустая (rowspan=2), в Excel будет "№"
                    pass
                elif cell_index == 1 and rowspan > 1:
                    # Вторая ячейка - "Аты-жөні" (rowspan=2)
                    headers_data["second_col_name"] = text if text else "Аты-жөні"
                else:
                    # Все остальные ячейки первой строки - предметы (включая пустые!)
                    # Важно: добавляем ВСЕ предметы, даже с пустым текстом, чтобы сохранить правильное количество колонок
                    headers_data["first_row"].append({
                        "text": text if text else "",  # Пустой текст для пустых предметов
                        "colspan": colspan
                    })
                    # Добавляем предмет в список столько раз, сколько столбцов он занимает
                    for _ in range(colspan):
                        headers_data["subjects"].append(text if text else "")
                
                cell_index += 1
            
            # --- Вторая строка заголовков: четверти ---
            second_row = rows[1]
            second_cells = second_row.find_elements(By.TAG_NAME, "th")
            
            print(f"  - Ячеек во второй строке заголовков: {len(second_cells)}")
            
            # Извлекаем последовательность четвертей из второй строки
            # Первые 2 ячейки пропущены из-за rowspan в первой строке
            raw_quarters: list[str] = []
            for cell in second_cells:
                text = cell.text.strip()
                if text:
                    raw_quarters.append(text)
            
            # Определяем уникальные четверти в порядке появления (обычно: І, ІІ, ІІІ, ІV, Ж)
            base_quarters: list[str] = []
            for q in raw_quarters:
                if q not in base_quarters:
                    base_quarters.append(q)
            
            # Если не смогли определить из таблицы — используем стандартный набор
            if not base_quarters:
                base_quarters = ["І", "ІІ", "ІІІ", "ІV", "Ж"]
            
            # Строим вторую строку заголовков: для каждого предмета повторяем базовую последовательность четвертей
            headers_data["second_row"] = []
            for header in headers_data["first_row"]:
                colspan = header.get("colspan", 1)
                
                # Для каждого предмета добавляем последовательность четвертей (І, ІІ, ІІІ, ІV, Ж)
                for q in base_quarters:
                    headers_data["second_row"].append(q)
            
            # Сохраняем информационные поля про четверти
            headers_data["unique_quarters"] = base_quarters
            headers_data["quarters_per_subject"] = len(base_quarters)
            
            # Отладочная информация
            print("✓ Извлечено заголовков:")
            print(f"  - Первая колонка: '{headers_data['first_col_name']}'")
            print(f"  - Вторая колонка: '{headers_data['second_col_name']}'")
            print(f"  - Предметов в первой строке: {len(headers_data['first_row'])}")
            print(f"  - Четвертей во второй строке (всего столбцов по предметам): {len(headers_data['second_row'])}")
            if headers_data["first_row"]:
                # Показываем все предметы, включая пустые
                subjects_list = [h['text'] if h['text'] else '(пустой)' for h in headers_data['first_row']]
                print(f"  - Все предметы (первые 10): {subjects_list[:10]}")
                # Проверяем наличие пустых предметов
                empty_count = sum(1 for h in headers_data['first_row'] if not h['text'])
                if empty_count > 0:
                    print(f"  - ⚠ Найдено пустых предметов: {empty_count}")
            if headers_data["second_row"]:
                print(f"  - Примеры четвертей: {headers_data['second_row'][:10]}")
            
            return headers_data
            
        except Exception as e:
            print(f"⚠ Ошибка при извлечении заголовков: {e}")
            import traceback
            traceback.print_exc()
            return {
                "first_row": [],
                "second_row": [],
                "subjects": [],
                "first_col_name": "№",
                "second_col_name": "Аты-жөні"
            }
    
    def _extract_table_body(self, table):
        """Извлечение данных из тела таблицы 'Сапа' из модального окна.
        
        В HTML структура tbody (из сапа.html):
        - Строки с данными учеников:
          - Первая колонка: порядковый номер (1, 2, 3...) - берем как есть
          - Вторая колонка: ФИО (например, "Байсеитова Аяна") - берем как есть
          - Остальные колонки: оценки по предметам - берем как есть
        - Служебные строки в конце (пропускаем):
          - Строка с colspan="999" (пустая разделительная)
          - Строки SUM11 (badge-success, badge-info, badge-warning) с colspan="2" и текстом "5", "4", "3"
          - Строки "үлгерімі" и "сапасы" с colspan="2"
        
        В Excel структура:
        - Первая колонка: порядковый номер (№) - берем из первой колонки HTML
        - Вторая колонка: ФИО (Аты-жөні) - берем из второй колонки HTML
        - Остальные колонки: данные по предметам - берем начиная с 3-й колонки HTML
        """
        try:
            table_data = []
            
            # Ищем tbody
            tbody = table.find_elements(By.TAG_NAME, "tbody")
            if not tbody:
                return table_data
            
            rows = tbody[0].find_elements(By.TAG_NAME, "tr")
            
            for row_idx, row in enumerate(rows):
                cells = row.find_elements(By.TAG_NAME, "td")
                
                if not cells or len(cells) < 2:
                    continue
                
                # Пропускаем служебные строки
                # 1. Строки с colspan="999" (разделительная пустая строка)
                try:
                    first_cell_colspan = cells[0].get_attribute("colspan")
                    if first_cell_colspan and int(first_cell_colspan) >= 999:
                        continue
                except:
                    pass
                
                # 2. Строки SUM11 (имеют классы badge-success, badge-info, badge-warning, badge-danger)
                row_class = row.get_attribute("class") or ""
                if any(badge in row_class for badge in ["badge-success", "badge-info", "badge-warning", "badge-danger"]):
                    # Проверяем, если первая ячейка содержит "5", "4", "3", "2" в жирном шрифте - это строка SUM11
                    first_cell_text = cells[0].text.strip() if cells else ""
                    if first_cell_text in ["5", "4", "3", "2"]:
                        try:
                            bold_elements = cells[0].find_elements(By.TAG_NAME, "b")
                            if bold_elements:
                                continue
                        except:
                            pass
                
                # 3. Строки "үлгерімі" и "сапасы"
                first_cell_text = cells[0].text.strip() if cells else ""
                if first_cell_text in ["үлгерімі", "сапасы"]:
                    continue
                
                # Извлекаем данные ученика
                row_data = []
                
                # Первая колонка - порядковый номер (берем как есть)
                if len(cells) > 0:
                    num_text = cells[0].text.strip()
                    row_data.append(num_text)
                
                # Вторая колонка - ФИО (берем как есть)
                if len(cells) > 1:
                    fio_text = cells[1].text.strip()
                    row_data.append(fio_text)
                
                # Остальные колонки - данные по предметам (берем как есть)
                for cell_idx in range(2, len(cells)):
                    cell = cells[cell_idx]
                    text = cell.text.strip()
                    # Очищаем текст от лишних пробелов и переносов
                    text = re.sub(r'\s+', ' ', text)
                    row_data.append(text)
                
                # Добавляем строку только если есть данные
                if row_data and len(row_data) >= 2:
                    # Проверяем, что первая колонка - это число (номер), а вторая - не пустая (ФИО)
                    try:
                        num_value = int(row_data[0]) if row_data[0] else None
                        fio_value = row_data[1] if len(row_data) > 1 else ""
                        if num_value and fio_value and fio_value.strip():
                            table_data.append(row_data)
                    except (ValueError, IndexError):
                        # Если первая колонка не число, но есть ФИО - тоже добавляем
                        if len(row_data) > 1 and row_data[1] and row_data[1].strip():
                            table_data.append(row_data)
            
            print(f"✓ Извлечено строк данных: {len(table_data)}")
            if table_data:
                print(f"  - Колонок в первой строке: {len(table_data[0])}")
                if len(table_data[0]) > 1:
                    print(f"  - Пример номера (1-я колонка): {table_data[0][0] if len(table_data[0]) > 0 else 'N/A'}")
                    print(f"  - Пример ФИО (2-я колонка): {table_data[0][1] if len(table_data[0]) > 1 else 'N/A'}")
            
            return table_data
            
        except Exception as e:
            print(f"⚠ Ошибка при извлечении данных: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def save_to_excel(self, table_data, class_name, output_file="success_data.xlsx"):
        """Сохранение данных в Excel с сохранением структуры таблицы
        Добавляет новый лист в существующий файл или создает новый файл
        """
        try:
            print(f"\nСохранение данных в Excel: {output_file}")
            
            # Очищаем название класса для использования в имени листа
            sheet_name = class_name.replace("«", "").replace("»", "").replace('"', "").replace("/", "_")
            if len(sheet_name) > 31:  # Ограничение Excel
                sheet_name = sheet_name[:31]
            
            # Загружаем существующий файл или создаем новый
            if os.path.exists(output_file):
                try:
                    wb = load_workbook(output_file)
                    # Проверяем, существует ли уже лист с таким именем
                    if sheet_name in wb.sheetnames:
                        # Удаляем старый лист
                        wb.remove(wb[sheet_name])
                        print(f"⚠ Лист '{sheet_name}' уже существует, будет перезаписан")
                except Exception as e:
                    print(f"⚠ Не удалось загрузить существующий файл: {e}. Создаю новый.")
                    wb = Workbook()
                    if wb.active:
                        wb.remove(wb.active)
            else:
                wb = Workbook()
                if wb.active:
                    wb.remove(wb.active)
            
            # Создаем новый лист для класса
            ws = wb.create_sheet(title=sheet_name)
            
            if not table_data or not table_data.get("headers") or not table_data.get("data"):
                print("✗ Нет данных для сохранения")
                return False
            
            headers = table_data["headers"]
            data = table_data["data"]
            
            # Стили для заголовков
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Записываем заголовки
            current_row = 1
            first_col_name = headers.get("first_col_name", "")  # Первая колонка без заголовка
            second_col_name = headers.get("second_col_name", "Аты-жөні")  # Вторая колонка - ФИО
            
            # Первая строка заголовков
            col = 1
            # Первая колонка - пустая (будет объединена с rowspan=2)
            ws.cell(row=current_row, column=col, value=first_col_name)
            ws.cell(row=current_row, column=col).font = header_font
            ws.cell(row=current_row, column=col).fill = header_fill
            ws.cell(row=current_row, column=col).alignment = header_alignment
            ws.cell(row=current_row, column=col).border = border
            col += 1
            
            # Вторая колонка - "Аты-жөні" (будет объединена с rowspan=2)
            ws.cell(row=current_row, column=col, value=second_col_name)
            ws.cell(row=current_row, column=col).font = header_font
            ws.cell(row=current_row, column=col).fill = header_fill
            ws.cell(row=current_row, column=col).alignment = header_alignment
            ws.cell(row=current_row, column=col).border = border
            col += 1
            
            # Предметы с colspan (включая пустые!)
            if headers.get("first_row"):
                for header in headers["first_row"]:
                    start_col = col
                    end_col = col + header["colspan"] - 1
                    
                    # Объединяем ячейки для предмета (даже если он пустой)
                    if header["colspan"] > 1:
                        ws.merge_cells(start_row=current_row, start_column=start_col,
                                     end_row=current_row, end_column=end_col)
                    
                    # Записываем текст предмета (может быть пустым)
                    cell = ws.cell(row=current_row, column=start_col, value=header["text"] if header["text"] else "")
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                    
                    col = end_col + 1
            
            # Вторая строка заголовков (четверти)
            current_row = 2
            col = 1
            # Первая колонка - пустая (уже записана, будет объединена)
            col += 1
            # Вторая колонка - "Аты-жөні" (уже записана, будет объединена)
            col += 1
            
            # Записываем четверти для каждого предмета
            all_quarters = headers.get("second_row", [])
            
            # Проверяем соответствие количества предметов и четвертей
            total_subject_cols = sum(h.get("colspan", 1) for h in headers.get("first_row", []))
            if len(all_quarters) != total_subject_cols:
                print(f"⚠ Несоответствие: предметов (с учетом colspan): {total_subject_cols}, четвертей: {len(all_quarters)}")
            
            for quarter in all_quarters:
                cell = ws.cell(row=current_row, column=col, value=quarter)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                col += 1
            
            # Объединяем первую колонку для двух строк заголовков (пустая)
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            first_cell = ws.cell(row=1, column=1)
            first_cell.value = first_col_name
            first_cell.font = header_font
            first_cell.fill = header_fill
            first_cell.alignment = header_alignment
            first_cell.border = border
            
            # Объединяем вторую колонку для двух строк заголовков ("Аты-жөні")
            ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
            second_cell = ws.cell(row=1, column=2)
            second_cell.value = second_col_name
            second_cell.font = header_font
            second_cell.fill = header_fill
            second_cell.alignment = header_alignment
            second_cell.border = border
            
            # Записываем данные
            current_row = 3
            
            for row_data in data:
                col = 1
                
                # Первая колонка - порядковый номер (№) из первой колонки row_data
                if len(row_data) > 0:
                    num_value = row_data[0]
                    cell = ws.cell(row=current_row, column=col, value=num_value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    col += 1
                else:
                    col += 1
                
                # Вторая колонка - ФИО (вторая колонка в row_data)
                if len(row_data) > 1:
                    fio_value = row_data[1]
                    cell = ws.cell(row=current_row, column=col, value=fio_value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border
                    col += 1
                else:
                    col += 1
                    
                # Остальные колонки - данные по предметам (начиная с 3-й колонки в row_data)
                for data_idx in range(2, len(row_data)):
                    cell_value = row_data[data_idx]
                    cell = ws.cell(row=current_row, column=col, value=cell_value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    col += 1
                
                current_row += 1
            
            # Настраиваем ширину колонок
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in ws[column_letter]:
                    try:
                        if row.value:
                            max_length = max(max_length, len(str(row.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            wb.save(output_file)
            print(f"✓ Данные сохранены в файл: {output_file}")
            print(f"  Лист: {sheet_name}")
            print(f"  Строк данных: {len(data)}")
            
            return True
            
        except Exception as e:
            print(f"✗ Ошибка при сохранении в Excel: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """Основная функция для запуска парсера"""
    scraper = MektepScraper()
    scraper.setup_driver()
    
    # Авторизация
    if not scraper.login():
        print("✗ Не удалось авторизоваться. Завершение работы.")
        if scraper.driver:
            scraper.driver.quit()
        return
    
    # Переход на страницу отчетов
    if not scraper.navigate_to_reports():
        print("✗ Не удалось перейти на страницу отчетов. Завершение работы.")
        if scraper.driver:
            scraper.driver.quit()
        return
    
    # Выбор школы
    if not scraper.select_school():
        print("✗ Не удалось выбрать школу. Завершение работы.")
        if scraper.driver:
            scraper.driver.quit()
        return
    
    # Получение списка классов (вкладок)
    classes = scraper.get_classes_list()
    if not classes:
        print("✗ Классы не найдены. Завершение работы.")
        if scraper.driver:
            scraper.driver.quit()
        return
    
    # Выводим список классов (вкладок)
    print(f"\n{'='*60}")
    print("СПИСОК КЛАССОВ (ВКЛАДКИ):")
    print(f"{'='*60}")
    for cls in classes:
        active_mark = " (активен)" if cls['is_active'] else ""
        print(f"{cls['index']}. {cls['name']}{active_mark}")
    print(f"{'='*60}\n")
    
    # Выбор класса (вкладки)
    while True:
        try:
            choice = input(f"Выберите номер класса (1-{len(classes)}): ").strip()
            class_index = int(choice)
            if 1 <= class_index <= len(classes):
                selected_class = classes[class_index - 1]
                break
            else:
                print(f"Пожалуйста, введите число от 1 до {len(classes)}")
        except ValueError:
            print("Пожалуйста, введите корректный номер")
        except KeyboardInterrupt:
            print("\nОтменено пользователем")
            if scraper.driver:
                scraper.driver.quit()
            return
    
    # Выбираем вкладку класса
    if not scraper.select_class_tab(selected_class['number']):
        print("✗ Не удалось выбрать класс. Завершение работы.")
        if scraper.driver:
            scraper.driver.quit()
        return
    
    # Получаем список классов из таблицы
    class_groups = scraper.get_class_groups_from_table()
    if not class_groups:
        print(f"✗ Классы {selected_class['number']} класса не найдены в таблице. Завершение работы.")
        if scraper.driver:
            scraper.driver.quit()
        return
    
    # Выводим список классов
    print(f"\n{'='*60}")
    print(f"СПИСОК КЛАССОВ {selected_class['number']} КЛАССА:")
    print(f"{'='*60}")
    for group in class_groups:
        print(f"{group['index']}. {group['name']} (Литера: {group['letter']}, Тип: {group['type']}, "
              f"Язык: {group['language']}, Смена: {group['shift']}, "
              f"Учащиеся: {group['students']}, Руководитель: {group['teacher']})")
    print(f"{'='*60}\n")
    
    # Спрашиваем, обрабатывать все классы или один
    while True:
        try:
            choice = input("Обработать все классы из параллели? (да/нет): ").strip().lower()
            if choice in ['да', 'д', 'yes', 'y']:
                process_all = True
                break
            elif choice in ['нет', 'н', 'no', 'n']:
                process_all = False
                break
            else:
                print("Пожалуйста, введите 'да' или 'нет'")
        except KeyboardInterrupt:
            print("\nОтменено пользователем")
            if scraper.driver:
                scraper.driver.quit()
            return
    
    output_file = "success_data.xlsx"
    
    if process_all:
        # Обрабатываем все классы из параллели
        print(f"\n{'='*60}")
        print(f"ОБРАБОТКА ВСЕХ КЛАССОВ {selected_class['number']} КЛАССА")
        print(f"{'='*60}\n")
        
        for group in class_groups:
            print(f"\n{'='*60}")
            print(f"Обработка класса: {group['name']}")
            print(f"{'='*60}")
            
            # Кликаем по кнопке "Успеваемость" для текущего класса
            if not group.get('button'):
                print(f"⚠ Кнопка 'Успеваемость' не найдена для {group['name']}, пропускаем")
                continue
            
            try:
                # Прокручиваем к кнопке
                scraper.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", group['button'])
                time.sleep(0.5)
                
                # Используем JavaScript для клика
                scraper.driver.execute_script("arguments[0].click();", group['button'])
                print(f"✓ Клик по кнопке 'Успеваемость' для {group['name']}")
                
                # Ждем открытия модального окна
                time.sleep(1)
                
                # Ждем появления и открытия модального окна
                try:
                    scraper.wait.until(
                        EC.any_of(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa.modal.show")),
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa.modal:not(.fade)")),
                            EC.presence_of_element_located((By.ID, "classSapa"))
                        )
                    )
                    time.sleep(2)
                    
                    # Ждем появления таблицы
                    try:
                        scraper.wait.until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#classSapa table"))
                        )
                        print(f"✓ Модальное окно 'Сапа' открыто для {group['name']}")
                    except TimeoutException:
                        print(f"⚠ Таблица не появилась в модальном окне для {group['name']}, но продолжаем...")
                except TimeoutException:
                    print(f"⚠ Модальное окно не найдено для {group['name']}, пропускаем")
                    continue
                
                # Извлечение данных из модального окна
                table_data = scraper.extract_modal_table_data()
                if not table_data:
                    print(f"✗ Не удалось извлечь данные для {group['name']}, пропускаем")
                    scraper.close_modal()
                    continue
                
                # Сохранение данных в Excel
                class_name = group['name']
                if scraper.save_to_excel(table_data, class_name, output_file):
                    print(f"✓ Данные для {class_name} сохранены в файл: {output_file}")
                else:
                    print(f"✗ Не удалось сохранить данные для {class_name}")
                
                # Закрываем модальное окно
                scraper.close_modal()
                time.sleep(0.5)  # Небольшая пауза перед следующим классом
                
            except Exception as e:
                print(f"✗ Ошибка при обработке {group['name']}: {e}")
                scraper.close_modal()
                continue
        
        print(f"\n{'='*60}")
        print(f"ОБРАБОТКА ЗАВЕРШЕНА")
        print(f"{'='*60}")
        print(f"✓ Все данные сохранены в файл: {output_file}")
        
    else:
        # Обрабатываем только один выбранный класс
        selected_group = scraper.select_class_group(class_groups)
        if not selected_group:
            print("✗ Не удалось выбрать класс. Завершение работы.")
            if scraper.driver:
                scraper.driver.quit()
            return
        
        # Извлечение данных из модального окна
        table_data = scraper.extract_modal_table_data()
        if not table_data:
            print("✗ Не удалось извлечь данные из модального окна. Завершение работы.")
            if scraper.driver:
                scraper.driver.quit()
            return
        
        # Сохранение данных в Excel
        class_name = selected_group['name']
        if scraper.save_to_excel(table_data, class_name, output_file):
            print(f"\n✓ Данные успешно сохранены в файл: {output_file}")
        else:
            print("✗ Не удалось сохранить данные в Excel")
    
    # Закрываем браузер
    if scraper.driver:
        scraper.driver.quit()


if __name__ == "__main__":
    main()